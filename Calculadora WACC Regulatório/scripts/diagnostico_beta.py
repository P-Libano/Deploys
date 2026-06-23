"""
Diagnóstico completo da trilha de cálculo do Beta — C1 vs planilha ANEEL.
Mostra cada dado, de onde vem, e como chega no beta_l = 0.769238.

Executar: python scripts/diagnostico_beta.py
"""
import pandas as pd
import openpyxl

SEP = "=" * 72

def main():
    print(SEP)
    print("TRILHA COMPLETA DO CALCULO DE BETA - C1 vs PLANILHA ANEEL")
    print(SEP)

    # --- 1. Fixture ---
    df = pd.read_csv("wacc_regulatorio/data/fixtures/beta_historico.csv", sep=";")
    df = df.sort_values("ano").reset_index(drop=True)

    print("\n--- 1. FIXTURE beta_historico.csv  (extraído de WACC Histórico linha 22) ---")
    print(df[["ano", "beta_u_eua", "dv_brasil", "ev_brasil", "T_brasil", "beta_l_brasil"]].to_string(index=False))
    print()
    print("  Cada linha = uma janela OLS de 5 anos (Oct-Sep) de utilities americanas.")
    print("  beta_u_eua  = beta desalavancado (D/E americano de cada empresa)")
    print("  beta_l_brasil = beta_u_eua re-alavancado com D/E brasileiro DO ANO")
    print("                = beta_u_eua * (1 + (1 - T_brasil) * (dv_brasil / ev_brasil))")
    print()
    for _, r in df.iterrows():
        de_br = r.dv_brasil / r.ev_brasil
        check = r.beta_u_eua * (1 + (1 - r.T_brasil) * de_br)
        print(f"  {int(r.ano)}: {r.beta_u_eua:.6f} * (1 + {1-r.T_brasil:.4f} * {de_br:.4f}) = {check:.6f}  (fixture={r.beta_l_brasil:.6f}  diff={abs(check-r.beta_l_brasil)*1e6:.1f}µbp)")

    # --- 2. Ler beta_l da aba WACC Histórico no xlsx ---
    print()
    print("--- 2. XLSX aba 'WACC Histórico', linha 6 (Beta Alavancado por ano) ---")
    wb = openpyxl.load_workbook(
        "anexo-despacho-1174-2026-aneel-2-Anexo_Memoria_de_Calculo_WACC_2026.xlsx",
        data_only=True,
    )
    ws_h = wb["WACC Histórico"]
    rows_h = list(ws_h.iter_rows(values_only=True))
    row3 = rows_h[2]   # cabeçalho de anos
    row6 = rows_h[5]   # Beta Alavancado

    beta_l_xlsx = {}
    for i in range(2, 17):
        if row3[i] is not None and isinstance(row3[i], (int, float)) and row6[i] is not None:
            beta_l_xlsx[int(row3[i])] = float(row6[i])

    for ano, val in sorted(beta_l_xlsx.items()):
        fv = df[df["ano"] == ano]["beta_l_brasil"].values
        match = ""
        if len(fv):
            diff = abs(val - fv[0])
            match = f"  == fixture: {diff:.2e}"
        print(f"  {ano}: {val:.6f}{match}")

    # --- 3. Fórmula no xlsx ---
    print()
    print("--- 3. FORMULA na aba 'WACC para aplicação', linha 7 (Beta Alavancado) ---")
    wb_f = openpyxl.load_workbook(
        "anexo-despacho-1174-2026-aneel-2-Anexo_Memoria_de_Calculo_WACC_2026.xlsx",
        data_only=False,
    )
    ws_wacc = wb_f["WACC para aplicação"]
    rows_wacc = list(ws_wacc.iter_rows(values_only=False))
    row7 = rows_wacc[6]
    print(f"  Coluna 2026 (beta_l final): {row7[10].value}")
    print(f"  Coluna 2025:                {row7[9].value}")
    print(f"  Coluna 2024:                {row7[8].value}")
    print()
    print("  Interpretação:")
    print("  K6:O6  =  anos 2021, 2022, 2023, 2024, 2025  (cols K..O do WACC Histórico)")
    print("  =AVERAGE(...)  =  média simples das 5 mais recentes de beta_l_brasil")

    # --- 4. Cálculo passo a passo ---
    print()
    print("--- 4. CALCULO PASSO A PASSO (codigo calcular_beta_from_historico) ---")
    ultimas5 = df[df["ano"].isin([2021, 2022, 2023, 2024, 2025])]
    vals = ultimas5["beta_l_brasil"].values
    soma = vals.sum()
    media = vals.mean()

    for i, (_, r) in enumerate(ultimas5.iterrows()):
        print(f"  {int(r.ano)}: {r.beta_l_brasil:.6f}")
    print(f"  -------")
    print(f"  Soma  = {soma:.6f}")
    print(f"  Media = {soma:.6f} / 5 = {media:.6f}")
    print()
    ref = 0.769238691706201
    print(f"  Ref ANEEL (xlsx)  = {ref:.6f}")
    print(f"  Nosso calculo     = {media:.6f}")
    print(f"  Delta             = {(media - ref) * 10000:+.3f}bp")

    # --- 5. E/V e D/V ---
    print()
    print("--- 5. ESTRUTURA DE CAPITAL (ano mais recente = 2025) ---")
    rec = df[df["ano"] == 2025].iloc[0]
    print(f"  E/V = {rec.ev_brasil:.6f} ({rec.ev_brasil*100:.4f}%)  [ref ANEEL: 60.2261%]")
    print(f"  D/V = {rec.dv_brasil:.6f} ({rec.dv_brasil*100:.4f}%)  [ref ANEEL: 39.7739%]")

    # --- 6. Resultado final no WACC ---
    print()
    print("--- 6. COMO BETA ENTRA NO WACC ---")
    rf = 0.051362
    erp = 0.068481
    ke = rf + media * erp
    print(f"  Ke = Rf + beta_l * ERP")
    print(f"     = {rf:.6f} + {media:.6f} * {erp:.6f}")
    print(f"     = {ke:.6f}  ({ke*100:.4f}%)")
    print(f"     Ref ANEEL Ke = 10.4055%")
    print()
    print("  Nota: beta_u NÃO entra no Ke. É informação diagnóstica apenas.")
    print(SEP)


if __name__ == "__main__":
    main()
