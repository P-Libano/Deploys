"""
Extrai CSVs limpos da planilha ANEEL para wacc_regulatorio/data/fixtures/.

Execute uma vez antes de usar o módulo:
    python scripts/extrair_fixtures.py

Nenhum módulo de cálculo deve importar openpyxl diretamente.
"""
import os
import sys
import warnings
import datetime
from pathlib import Path

import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

XLSX = Path(__file__).parent.parent / "anexo-despacho-1174-2026-aneel-2-Anexo_Memoria_de_Calculo_WACC_2026.xlsx"
OUT = Path(__file__).parent.parent / "wacc_regulatorio" / "data" / "fixtures"
OUT.mkdir(parents=True, exist_ok=True)

SEP = ";"


def _wb():
    return openpyxl.load_workbook(XLSX, read_only=True, data_only=True)


# ---------------------------------------------------------------------------
# NTN-B diário
# ---------------------------------------------------------------------------
def extrair_ntnb():
    print("-> Extraindo NTN-B diario...")
    wb = _wb()
    ws = wb["NTN-B "]

    # Row 3: col 1 = "Vencimento", depois vencimentos em colunas pares (2,4,6,...)
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    # Monta mapa: coluna (0-based) -> data de vencimento
    venc_cols = {}
    for i, v in enumerate(row3):
        if isinstance(v, datetime.datetime):
            venc_cols[i] = v.date()  # coluna 0-based

    # Dados: row 5 em diante
    # Datas podem ser datetime (até 2013) ou string "DD/MM/YYYY" (a partir de 2014)
    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        data_val = row[0]
        if data_val is None:
            continue
        if isinstance(data_val, datetime.datetime):
            data = data_val.date()
        elif isinstance(data_val, str):
            try:
                data = datetime.datetime.strptime(data_val.strip(), "%d/%m/%Y").date()
            except ValueError:
                continue
        else:
            continue  # linha não-dado (ex: rótulos de médias anuais)
        for col0, venc in venc_cols.items():
            compra = row[col0]           # taxa compra manhã
            venda = row[col0 + 1] if col0 + 1 < len(row) else None  # taxa venda manhã
            if compra is not None and isinstance(compra, (int, float)):
                rec = {
                    "data": data,
                    "vencimento": venc,
                    "taxa_compra_manha": float(compra),
                    "taxa_venda_manha": float(venda) if isinstance(venda, (int, float)) else None,
                }
                records.append(rec)

    wb.close()
    df = pd.DataFrame(records)
    df["data"] = pd.to_datetime(df["data"])
    df["vencimento"] = pd.to_datetime(df["vencimento"])
    df.to_csv(OUT / "ntnb_diario.csv", sep=SEP, index=False, encoding="utf-8")
    n_venda = df["taxa_venda_manha"].notna().sum()
    print(f"   {len(df):,} linhas -> ntnb_diario.csv  (taxa_venda: {n_venda:,} obs)")


# ---------------------------------------------------------------------------
# PRM — S&P500 mensal + Treasury 10Y
# ---------------------------------------------------------------------------
def extrair_prm():
    print("-> Extraindo PRM (S&P500 + Treasury 10Y)...")
    wb = _wb()
    ws = wb["PRM"]

    # Row 4: Data, S&P500, (blank), Rm, (blank), Rf, PRM
    # Cols: 1=data, 2=sp500, 4=rm, 6=rf, 7=prm (1-based)
    # Row 5 onward = data

    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        data_val = row[0]
        sp500 = row[1]
        rf = row[5]  # col 6 (0-based = 5)
        if not isinstance(data_val, datetime.datetime):
            continue
        if sp500 is None:
            continue
        records.append({
            "data": data_val.date(),
            "sp500": float(sp500) if sp500 is not None else None,
            "rf_tbill": float(rf) if isinstance(rf, (int, float)) else None,
        })

    wb.close()
    df = pd.DataFrame(records)
    df["data"] = pd.to_datetime(df["data"])
    df.to_csv(OUT / "prm_sp500.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> prm_sp500.csv")


# ---------------------------------------------------------------------------
# EMBI — série diária do IPEADATA (aba OE)
# ---------------------------------------------------------------------------
def extrair_embi():
    print("-> Extraindo EMBI diário (aba OE)...")
    wb = _wb()
    ws = wb["OE"]

    # EMBI data: cols 12-14 (1-based) = date, bps, decimal
    # Data starts at row 6 (row 4 = header, row 5 = URL)
    records = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        data_val = row[11]   # col 12 (0-based = 11)
        bps = row[12]        # col 13
        decimal = row[13]    # col 14
        if not isinstance(data_val, datetime.datetime):
            continue
        if decimal is None:
            continue
        records.append({
            "data": data_val.date(),
            "embi_bps": int(bps) if isinstance(bps, (int, float)) else None,
            "embi_decimal": float(decimal),
        })

    wb.close()
    df = pd.DataFrame(records)
    df["data"] = pd.to_datetime(df["data"])
    df.to_csv(OUT / "embi_diario.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> embi_diario.csv")


# ---------------------------------------------------------------------------
# Debêntures
# ---------------------------------------------------------------------------
def extrair_debentures():
    print("-> Extraindo Debêntures...")
    wb = _wb()
    ws = wb["Debentures "]

    # Row 3 = headers (1-based), data rows 4-669
    # Colunas relevantes (1-based):
    # 1=código, 2=area, 3=índice, 4=empresa, 5=data_emissao, 6=ano,
    # 7=data_vencimento, 8=qtd, 9=valor_nominal, 10=rentabilidade,
    # 12=juros, 14=anos, 15=dias_corridos, 16=DI, 17=swap_di_ipca,
    # 18=taxa_nominal, 19=inflacao, 20=taxa_real

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        codigo = row[0]
        if codigo is None:
            continue
        area = row[1]
        indice = row[2]
        empresa = row[3]
        data_emissao = row[4]
        ano = row[5]
        data_venc = row[6]
        qtd = row[7]
        vnom = row[8]
        rentab = row[9]
        juros = row[11]   # col 12 (0-based=11)
        anos = row[13]    # col 14
        dias = row[14]    # col 15
        di = row[15]      # col 16 — "Sem Dados" para antigas
        swap = row[16]    # col 17
        taxa_nom = row[17]  # col 18
        inflacao = row[18]  # col 19
        taxa_real = row[19]  # col 20

        records.append({
            "codigo": str(codigo).strip(),
            "area": str(area).strip() if area else None,
            "indice": str(indice).strip() if indice else None,
            "empresa": str(empresa).strip() if empresa else None,
            "data_emissao": data_emissao.date() if isinstance(data_emissao, datetime.datetime) else None,
            "ano": int(ano) if isinstance(ano, (int, float)) else None,
            "data_vencimento": data_venc.date() if isinstance(data_venc, datetime.datetime) else None,
            "quantidade": qtd,
            "valor_nominal": float(vnom) if isinstance(vnom, (int, float)) else None,
            "valor_emissao": float(qtd * vnom) if isinstance(qtd, (int, float)) and isinstance(vnom, (int, float)) else None,
            "rentabilidade_pct": float(rentab) if isinstance(rentab, (int, float)) else None,
            "juros_spread": float(juros) if isinstance(juros, (int, float)) else None,
            "anos": float(anos) if isinstance(anos, (int, float)) else None,
            "dias_corridos": int(dias) if isinstance(dias, (int, float)) else None,
            "di": float(di) if isinstance(di, (int, float)) else None,
            "swap_di_ipca": float(swap) if isinstance(swap, (int, float)) else None,
            "taxa_nominal_pct": float(taxa_nom) if isinstance(taxa_nom, (int, float)) else None,
            "inflacao_implicita": float(inflacao) if isinstance(inflacao, (int, float)) else None,
            "taxa_real": float(taxa_real) if isinstance(taxa_real, (int, float)) else None,
        })

    wb.close()
    df = pd.DataFrame(records)
    df["data_emissao"] = pd.to_datetime(df["data_emissao"])
    df["data_vencimento"] = pd.to_datetime(df["data_vencimento"])
    df.to_csv(OUT / "debentures.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> debentures.csv")


# ---------------------------------------------------------------------------
# Custo de Emissão
# ---------------------------------------------------------------------------
def extrair_custo_emissao():
    print("-> Extraindo Custo de Emissão...")
    wb = _wb()
    ws = wb["Custo de Emissao"]

    # Row 5 = headers, data rows 6+
    # Colunas relevantes (1-based):
    # 1=código, 2=empresa, 3=classificação, 9=data_emissao, 10=data_vencimento,
    # 11=qtd_emitida, 12=valor_nominal, 13=índice, 14=percentual_rent, 15=juros,
    # 16=juros_real, 18=valor_emissao, 19=custo_emissao, 20=prazo, 21=remuneracao_real

    records = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        codigo = row[0]
        if codigo is None:
            continue
        classificacao = row[2]
        # só aceita D, T, H (empresas do setor elétrico)
        data_emissao = row[8]   # col 9
        data_venc = row[9]      # col 10
        custo = row[18]         # col 19
        prazo = row[19]         # col 20
        rem_real = row[20]      # col 21

        if not isinstance(data_emissao, datetime.datetime):
            continue

        records.append({
            "codigo": str(codigo).strip(),
            "empresa": str(row[1]).strip() if row[1] else None,
            "classificacao": str(classificacao).strip() if classificacao else None,
            "data_emissao": data_emissao.date(),
            "data_vencimento": data_venc.date() if isinstance(data_venc, datetime.datetime) else None,
            "valor_nominal": float(row[11]) if isinstance(row[11], (int, float)) else None,
            "indice": str(row[12]).strip() if row[12] else None,
            "juros_real_pct": float(row[15]) if isinstance(row[15], (int, float)) else None,
            "valor_emissao_mi": float(row[17]) if isinstance(row[17], (int, float)) else None,
            "custo_emissao_pct": float(custo) if isinstance(custo, (int, float)) else None,
            "prazo_anos": float(prazo) if isinstance(prazo, (int, float)) else None,
            "remuneracao_real": float(rem_real) if isinstance(rem_real, (int, float)) else None,
        })

    wb.close()
    df = pd.DataFrame(records)
    df["data_emissao"] = pd.to_datetime(df["data_emissao"])
    df["data_vencimento"] = pd.to_datetime(df["data_vencimento"])
    df.to_csv(OUT / "custo_emissao.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> custo_emissao.csv")

    # Extrai mapeamento periodo → IPCA+DI (custo agregado da cesta por janela)
    # O ANEEL pré-computa este valor via fórmula no xlsx; não é derivável por média simples.
    wb2 = _wb()
    ws2 = wb2["Custo de Emissao"]
    periodos = {}
    for row in ws2.iter_rows(min_row=6, values_only=True):
        if row[0] is None:
            continue
        periodo = row[22]
        ipca_di = row[23]
        if periodo is not None and ipca_di is not None and isinstance(ipca_di, (int, float)):
            k = str(periodo).strip()
            if k not in periodos:
                periodos[k] = float(ipca_di)
    wb2.close()
    df_per = pd.DataFrame(
        [{"periodo": k, "custo_emissao_agregado": v} for k, v in sorted(periodos.items())]
    )
    df_per.to_csv(OUT / "custo_emissao_periodos.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df_per)} periodos -> custo_emissao_periodos.csv")


# ---------------------------------------------------------------------------
# WACC Histórico (2013-2025) — parâmetros por segmento
# ---------------------------------------------------------------------------
def extrair_wacc_historico():
    print("-> Extraindo WACC Histórico (2013-2025)...")
    wb = _wb()
    ws = wb["WACC Histórico"]

    # Layout (linhas, colunas 1-based):
    # Seção Transmissão: anos 2013-2025 nas colunas 3-15
    # Seção Beta e Estrutura Capital: anos 2013-2025 nas colunas 18-30
    # R5  = Rf
    # R6  = Beta_l
    # R7  = ERP
    # R8  = Business premium
    # R9  = Ke real antes impostos
    # R10 = Ke real depois impostos
    # R12 = Kd_debentures
    # R13 = Kd_custo_emissao
    # R14 = Kd real antes impostos
    # R15 = T
    # R16 = Kd real depois impostos
    # R18 = E/V (% Capital Próprio)
    # R19 = D/V (% Capital Terceiros)
    # R21 = WACC_di
    # R22 = WACC_ai
    # Seção direita (beta detalhado):
    # R18 (col 18-30) = Beta_u EUA desalavancado
    # R19 (col 18-30) = D/V estrutura Brasil
    # R22 (col 18-30) = Beta_l Brasil

    ANOS = list(range(2013, 2026))  # 13 anos
    COLS_LEFT = list(range(3, 16))   # 1-based cols 3-15 (0-based 2-14)
    COLS_RIGHT = list(range(18, 31)) # 1-based cols 18-30 (0-based 17-29)

    def _get_row(ws, row_num):
        return list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

    rows = {n: _get_row(ws, n) for n in range(3, 45)}

    records_transm = []
    for i, ano in enumerate(ANOS):
        c = COLS_LEFT[i] - 1  # 0-based
        cr = COLS_RIGHT[i] - 1

        def v(row_num, col):
            val = rows[row_num][col]
            return float(val) if isinstance(val, (int, float)) else None

        records_transm.append({
            "ano": ano,
            "segmento": "transmissao",
            "rf": v(5, c),
            "beta_l": v(6, c),
            "erp": v(7, c),
            "business_premium": v(8, c),
            "ke_real_ai": v(9, c),
            "ke_real_di": v(10, c),
            "kd_debentures": v(12, c),
            "kd_custo_emissao": v(13, c),
            "kd_real_ai": v(14, c),
            "T": v(15, c),
            "kd_real_di": v(16, c),
            "ev": v(18, c),
            "dv": v(19, c),
            "wacc_di": v(21, c),
            "wacc_ai": v(22, c),
            # Seção beta detalhado (lado direito)
            "beta_u": v(18, cr),
            "dv_br": v(19, cr),
            "beta_l_br": v(22, cr),
        })

    # Distribuição: começa na linha 27 (mesmo offset)
    # R29 = Rf, R30 = Beta_l, R31 = ERP, R32 = Business, R33 = Ke_ai,
    # R34 = Ke_di, R36 = Kd_deb, R37 = Kd_custo, R38 = Kd_ai,
    # R39 = T, R40 = Kd_di, R42 = E/V, R43 = D/V, R45 = WACC_di, R46 = WACC_ai
    # (Distribuição starts at row 27 with same structure offset by +24)

    wb.close()
    df = pd.DataFrame(records_transm)
    df.to_csv(OUT / "wacc_historico.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> wacc_historico.csv")


# ---------------------------------------------------------------------------
# WACC para aplicação (2018-2026) — valores oficiais por segmento
# ---------------------------------------------------------------------------
def extrair_wacc_aplicacao():
    print("-> Extraindo WACC para aplicação (2018-2026)...")
    wb = _wb()
    ws = wb["WACC para aplicação"]

    # Layout Transmissão: anos 2018-2026 nas colunas 3-11 (0-based 2-10)
    # R6  = Rf
    # R7  = Beta_l
    # R8  = ERP
    # R10 = Business premium
    # R11 = Ke_di
    # R13 = Kd_debentures
    # R14 = Kd_custo_emissao
    # R15 = Kd_real_ai
    # R16 = T
    # R17 = Kd_di
    # R19 = E/V
    # R20 = D/V
    # R22 = WACC_di
    # R23 = WACC_ai

    ANOS = list(range(2018, 2027))  # 2018-2026
    COLS = list(range(3, 12))       # 1-based 3-11 (0-based 2-10)

    def _get_row(ws, row_num):
        return list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

    rows = {n: _get_row(ws, n) for n in range(5, 25)}

    records = []
    for i, ano in enumerate(ANOS):
        c = COLS[i] - 1  # 0-based

        def v(row_num):
            val = rows[row_num][c]
            return float(val) if isinstance(val, (int, float)) else None

        records.append({
            "ano": ano,
            "segmento": "transmissao",
            "rf": v(6),
            "beta_l": v(7),
            "erp": v(8),
            "business_premium": v(10),
            "ke_real_di": v(11),
            "kd_debentures": v(13),
            "kd_custo_emissao": v(14),
            "kd_real_ai": v(15),
            "T": v(16),
            "kd_real_di": v(17),
            "ev": v(19),
            "dv": v(20),
            "wacc_di": v(22),
            "wacc_ai": v(23),
        })

    wb.close()
    df = pd.DataFrame(records)
    df.to_csv(OUT / "wacc_aplicacao.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> wacc_aplicacao.csv")
    # Show 2026 row as sanity check
    row_2026 = df[df["ano"] == 2026].iloc[0].to_dict()
    print(f"   2026 Transmissão: WACC_ai={row_2026['wacc_ai']:.4%}, WACC_di={row_2026['wacc_di']:.4%}")


# ---------------------------------------------------------------------------
# Beta — pré-calculados (do WACC Histórico, seção direita)
# ---------------------------------------------------------------------------
def extrair_beta_historico():
    """
    Extrai os betas pré-calculados do WACC Histórico (seção direita).
    Para Camada 2 ao vivo, o cálculo será refeito via yfinance.
    """
    print("-> Extraindo beta histórico pré-calculado...")
    # Já incluído em wacc_historico.csv (coluna beta_u)
    # Este arquivo adiciona o contexto da estrutura de capital americana
    wb = _wb()
    ws = wb["WACC Histórico"]

    ANOS = list(range(2013, 2026))
    COLS_RIGHT = list(range(18, 31))  # 1-based (0-based 17-29)

    def _get_row(ws, row_num):
        return list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

    rows = {n: _get_row(ws, n) for n in range(17, 23)}

    records = []
    for i, ano in enumerate(ANOS):
        cr = COLS_RIGHT[i] - 1  # 0-based

        def v(row_num):
            val = rows[row_num][cr]
            return float(val) if isinstance(val, (int, float)) else None

        records.append({
            "ano": ano,
            "beta_u_eua": v(18),
            "dv_brasil": v(19),
            "ev_brasil": 1.0 - (v(19) or 0),
            "T_brasil": v(20),
            "beta_l_brasil": v(22),
        })

    wb.close()
    df = pd.DataFrame(records)
    df.to_csv(OUT / "beta_historico.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> beta_historico.csv")


def extrair_embi_medias():
    """Extrai as médias anuais EMBI pré-calculadas pelo ANEEL (janelas 10 anos)."""
    print("-> Extraindo medias anuais EMBI (janelas pre-calculadas)...")
    wb = _wb()
    ws = wb["OE"]

    # Row 7 (0-based 6): labels de janela ex. '2008-2017', '2009-2018', ...
    # Row 8 (0-based 7): valores EMBI+ correspondentes
    # Cols 18-27 (0-based 17-26)
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]   # anos finais (2017-2025)
    row7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]   # labels janela
    row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]   # EMBI+ valor

    records = []
    for i in range(17, 27):  # 0-based cols 17-26
        ano_final = row6[i]
        janela = row7[i]
        embi_val = row8[i]
        if isinstance(embi_val, (int, float)) and isinstance(ano_final, (int, float)):
            records.append({
                "ano_wacc": int(ano_final) + 1,  # WACC publicado no ano seguinte
                "janela": str(janela),
                "embi_media_10a": float(embi_val),
            })

    wb.close()
    df = pd.DataFrame(records)
    df.to_csv(OUT / "embi_medias_anuais.csv", sep=SEP, index=False, encoding="utf-8")
    print(f"   {len(df):,} linhas -> embi_medias_anuais.csv")
    for _, r in df.iterrows():
        print(f"   WACC {r['ano_wacc']}: EMBI={r['embi_media_10a']:.5f} (janela {r['janela']})")


if __name__ == "__main__":
    print(f"\nPlanilha: {XLSX.name}")
    print(f"Destino:  {OUT}\n")

    if not XLSX.exists():
        print("ERRO: planilha não encontrada em", XLSX)
        sys.exit(1)

    extrair_ntnb()
    extrair_prm()
    extrair_embi()
    extrair_debentures()
    extrair_custo_emissao()
    extrair_wacc_historico()
    extrair_wacc_aplicacao()
    extrair_beta_historico()
    extrair_embi_medias()

    print("\nOK Todos os fixtures extraidos com sucesso.")
    print(f"  Local: {OUT}")
